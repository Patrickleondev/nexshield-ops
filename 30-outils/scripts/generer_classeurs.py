#!/usr/bin/env python3
"""Génère les classeurs de travail (.xlsx) à la charte de la société.

Usage : python3 30-outils/scripts/generer_classeurs.py [--out 90-templates/build]

Produit :
  - Classeur-mission.xlsx   : suivi complet d'une mission (à copier par mission)
  - Pilotage-societe.xlsx   : portefeuille, plan de charge, compétences
  - SoA-ISO27001.xlsx       : déclaration d'applicabilité

Les classeurs sont des GABARITS : on les copie dans le dossier de mission, on ne
les modifie pas ici. Toute évolution du gabarit passe par ce script, en PR.
"""
from __future__ import annotations
import argparse, os, sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import charte as C

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
# (les tableaux structures OOXML sont volontairement non utilises : voir tableau())

# --- Styles ------------------------------------------------------------------
F_TITRE   = Font(name=C.POLICE_TITRE, size=16, bold=True, color=C.PRIMAIRE)
F_SOUS    = Font(name=C.POLICE_CORPS, size=10, italic=True, color=C.TEXTE_FAIBLE)
F_ENTETE  = Font(name=C.POLICE_TITRE, size=10, bold=True, color=C.BLANC)
F_CORPS   = Font(name=C.POLICE_CORPS, size=10, color=C.TEXTE)
F_AIDE    = Font(name=C.POLICE_CORPS, size=9, italic=True, color=C.TEXTE_FAIBLE)
F_KPI     = Font(name=C.POLICE_TITRE, size=22, bold=True, color=C.PRIMAIRE)

R_ENTETE  = PatternFill("solid", fgColor=C.PRIMAIRE)
R_ALTERNE = PatternFill("solid", fgColor=C.FOND_ALTERNE)
R_ACCENT  = PatternFill("solid", fgColor=C.ACCENT)

_fin = Side(style="thin", color=C.BORDURE)
B_CELL = Border(left=_fin, right=_fin, top=_fin, bottom=_fin)

A_ENTETE = Alignment(horizontal="left", vertical="center", wrap_text=True)
A_CORPS  = Alignment(horizontal="left", vertical="top", wrap_text=True)
A_CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)


def feuille(wb: Workbook, titre: str, sous_titre: str, premiere: bool = False):
    ws = wb.active if premiere else wb.create_sheet()
    ws.title = titre
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{C.SOCIETE} - {titre}"
    ws["A1"].font = F_TITRE
    ws["A2"] = sous_titre
    ws["A2"].font = F_SOUS
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[3].height = 6
    return ws


def tableau(ws, ligne: int, colonnes: list[tuple[str, int]], nom: str,
            nb_lignes: int = 60, aide: str | None = None):
    """Pose un en-tête stylé + une zone de saisie, et rend la ligne de données."""
    if aide:
        ws.cell(ligne - 1, 1, aide).font = F_AIDE
    for i, (libelle, largeur) in enumerate(colonnes, start=1):
        c = ws.cell(ligne, i, libelle)
        c.font, c.fill, c.alignment, c.border = F_ENTETE, R_ENTETE, A_ENTETE, B_CELL
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.row_dimensions[ligne].height = 30

    for r in range(ligne + 1, ligne + 1 + nb_lignes):
        for i in range(1, len(colonnes) + 1):
            c = ws.cell(r, i)
            c.font, c.alignment, c.border = F_CORPS, A_CORPS, B_CELL
            if (r - ligne) % 2 == 0:
                c.fill = R_ALTERNE
        ws.row_dimensions[r].height = 22

    # Filtre automatique plutôt qu'un « tableau structuré » : même service au
    # quotidien, et lisible par tous les lecteurs (Excel, LibreOffice, aperçus
    # web). Les tableaux structurés d'OOXML cassent plusieurs visionneuses.
    ws.auto_filter.ref = f"A{ligne}:{get_column_letter(len(colonnes))}{ligne + nb_lignes}"
    ws.freeze_panes = ws.cell(ligne + 1, 1)
    return ligne + 1


def liste(ws, colonne: str, valeurs: list[str], debut: int, fin: int,
          message: str = ""):
    dv = DataValidation(type="list", formula1='"' + ",".join(valeurs) + '"',
                        allow_blank=True, showErrorMessage=True)
    dv.error = "Valeur hors liste. Voir CONVENTIONS.md."
    dv.errorTitle = "Valeur non autorisée"
    if message:
        dv.prompt, dv.promptTitle, dv.showInputMessage = message, "Aide", True
    ws.add_data_validation(dv)
    dv.add(f"{colonne}{debut}:{colonne}{fin}")


def couleurs_severite(ws, colonne: str, debut: int, fin: int):
    for nom, style in C.SEVERITES.items():
        ws.conditional_formatting.add(
            f"{colonne}{debut}:{colonne}{fin}",
            CellIsRule(operator="equal", formula=[f'"{nom}"'],
                       fill=PatternFill("solid", fgColor=style["fond"]),
                       font=Font(name=C.POLICE_CORPS, size=10, bold=True,
                                 color=style["texte"])))


# --- Classeur de mission -----------------------------------------------------
def classeur_mission(chemin: str) -> None:
    wb = Workbook()

    # Synthèse
    ws = feuille(wb, "Synthèse", "Tableau de bord - se remplit seul depuis les autres onglets", True)
    ws["A4"] = "Identification"
    ws["A4"].font = Font(name=C.POLICE_TITRE, size=12, bold=True, color=C.SECONDAIRE)
    champs = [("Client", ""), ("Mission", ""), ("Référence", "<CLIENT>-<type>-<nn>"),
              ("Chef de mission", ""), ("Doctrine appliquée", "pentest-audit v0.1"),
              ("Début", ""), ("Fin", ""), ("Statut", "cadrage"),
              ("Destruction des preuves", "= date de livraison + 90 j")]
    for i, (k, v) in enumerate(champs, start=5):
        ws.cell(i, 1, k).font = Font(name=C.POLICE_CORPS, size=10, bold=True, color=C.TEXTE)
        c = ws.cell(i, 2, v)
        c.font, c.border, c.fill = F_CORPS, B_CELL, R_ALTERNE
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42

    ws["D4"] = "Constatations par sévérité"
    ws["D4"].font = Font(name=C.POLICE_TITRE, size=12, bold=True, color=C.SECONDAIRE)
    for i, sev in enumerate(C.ORDRE_SEVERITES, start=5):
        c = ws.cell(i, 4, sev)
        c.font = Font(name=C.POLICE_CORPS, size=10, bold=True,
                      color=C.SEVERITES[sev]["texte"])
        c.fill = PatternFill("solid", fgColor=C.SEVERITES[sev]["fond"])
        c.alignment, c.border = A_ENTETE, B_CELL
        n = ws.cell(i, 5, f"=COUNTIF('Vulnérabilités'!D:D,\"{sev}\")")
        n.font, n.alignment, n.border = F_CORPS, A_CENTRE, B_CELL
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10

    ws["G4"] = "Avancement"
    ws["G4"].font = Font(name=C.POLICE_TITRE, size=12, bold=True, color=C.SECONDAIRE)
    ws["G5"] = "Tâches terminées"
    ws["G5"].font = F_CORPS
    ws["H5"] = "=COUNTIF(Tâches!F:F,\"Terminé\")&\" / \"&COUNTA(Tâches!A6:A200)"
    ws["H5"].font, ws["H5"].alignment = F_KPI, A_CENTRE
    ws["G7"] = "Jours consommés / vendus"
    ws["G7"].font = F_CORPS
    ws["H7"] = "=SUM(Tâches!H:H)&\" / \"&SUM('Plan de charge'!D:D)"
    ws["H7"].font, ws["H7"].alignment = F_KPI, A_CENTRE
    ws.column_dimensions["G"].width = 26
    ws.column_dimensions["H"].width = 18

    ws["A16"] = "Ce classeur ne contient AUCUNE preuve brute ni donnée client sensible."
    ws["A16"].font = Font(name=C.POLICE_CORPS, size=9, bold=True, color="991B1B")
    ws["A17"] = "Les preuves vivent au coffre chiffré - voir SECURITY.md §2."
    ws["A17"].font = F_AIDE

    # Vulnérabilités
    ws = feuille(wb, "Vulnérabilités", "Registre des constatations. Une ligne = une vulnérabilité.")
    cols = [("Identifiant", 18), ("Titre", 46), ("Actif affecté", 26), ("Sévérité", 14),
            ("CVSS v4.0", 11), ("Criticité métier", 16), ("CWE", 12),
            ("OWASP WSTG / ASVS", 20), ("MITRE ATT&CK / ATLAS", 20),
            ("Statut", 14), ("Découverte par", 16), ("Date", 12),
            ("Délai de correction", 18), ("Commentaire", 40)]
    d = tableau(ws, 5, cols, "Vulnerabilites", 80,
                aide="Identifiant : <CLIENT>-<AAAA>-<NNN>, continu par client et par année (CONVENTIONS.md).")
    liste(ws, "D", C.ORDRE_SEVERITES, d, d + 79,
          "Sévérité affichée. Si elle diverge du CVSS, justifier en commentaire.")
    liste(ws, "J", C.STATUTS_VULN, d, d + 79)
    liste(ws, "M", ["< 72 h", "30 jours", "90 jours", "Prochain cycle", "Aucune"], d, d + 79)
    couleurs_severite(ws, "D", d, d + 79)

    # Tâches
    ws = feuille(wb, "Tâches", "Répartition, échéances et charge. C'est l'outil de pilotage quotidien.")
    cols = [("#", 6), ("Phase PTES", 22), ("Tâche", 46), ("Responsable", 18),
            ("Appui", 16), ("Statut", 14), ("Priorité", 12),
            ("Charge estimée (j)", 15), ("Charge réelle (j)", 15),
            ("Début prévu", 13), ("Échéance", 13), ("Terminé le", 13),
            ("Bloqué par", 22), ("Notes", 36)]
    d = tableau(ws, 5, cols, "Taches", 80,
                aide="Une tâche sans responsable nommé et sans échéance n'existe pas. Règle non négociable.")
    liste(ws, "B", ["Pré-engagement", "Renseignement", "Modélisation de menaces",
                    "Analyse de vulnérabilités", "Exploitation", "Post-exploitation",
                    "Restitution", "Clôture"], d, d + 79)
    liste(ws, "F", ["À faire", "En cours", "Bloqué", "En revue", "Terminé"], d, d + 79)
    liste(ws, "G", ["Haute", "Moyenne", "Basse"], d, d + 79)
    ws.conditional_formatting.add(
        f"F{d}:F{d + 79}",
        CellIsRule(operator="equal", formula=['"Bloqué"'],
                   fill=PatternFill("solid", fgColor="991B1B"),
                   font=Font(name=C.POLICE_CORPS, size=10, bold=True, color=C.BLANC)))
    ws.conditional_formatting.add(
        f"F{d}:F{d + 79}",
        CellIsRule(operator="equal", formula=['"Terminé"'],
                   fill=PatternFill("solid", fgColor="15803D"),
                   font=Font(name=C.POLICE_CORPS, size=10, bold=True, color=C.BLANC)))
    # Échéance dépassée sur une tâche non terminée
    ws.conditional_formatting.add(
        f"K{d}:K{d + 79}",
        CellIsRule(operator="lessThan", formula=["TODAY()"],
                   fill=PatternFill("solid", fgColor="FEE2E2"),
                   font=Font(name=C.POLICE_CORPS, size=10, bold=True, color="991B1B")))

    # Plan de charge
    ws = feuille(wb, "Plan de charge", "Qui fait quoi, combien de jours, et à quel tarif.")
    cols = [("Membre", 22), ("Rôle sur la mission", 26), ("Séniorité", 14),
            ("Jours vendus", 13), ("Jours consommés", 15), ("Écart", 10),
            ("TJM", 12), ("Montant", 14), ("Disponibilité", 16), ("Commentaire", 36)]
    d = tableau(ws, 5, cols, "PlanDeCharge", 20)
    liste(ws, "B", ["Chef de mission", "Testeur", "Relecteur qualité",
                    "Référent juridique", "Appui technique"], d, d + 19)
    liste(ws, "C", ["Junior", "Confirmé", "Senior", "Expert"], d, d + 19)
    for r in range(d, d + 20):
        ws.cell(r, 6, f"=IF(D{r}=\"\",\"\",D{r}-E{r})").font = F_CORPS
        ws.cell(r, 8, f"=IF(D{r}=\"\",\"\",D{r}*G{r})").font = F_CORPS

    # Couverture
    ws = feuille(wb, "Couverture", "Annexe A du rapport. Trois statuts, et le motif est obligatoire si non exécuté.")
    cols = [("Identifiant", 20), ("Intitulé du test", 56), ("Applicable", 14),
            ("Statut", 18), ("Motif si non exécuté", 46), ("Testeur", 16)]
    d = tableau(ws, 5, cols, "Couverture", 130)
    liste(ws, "C", ["Oui", "Non"], d, d + 129)
    liste(ws, "D", ["Exécuté", "Non applicable", "Non exécuté"], d, d + 129)
    ws.conditional_formatting.add(
        f"D{d}:D{d + 129}",
        CellIsRule(operator="equal", formula=['"Non exécuté"'],
                   fill=PatternFill("solid", fgColor="FEF3C7"),
                   font=Font(name=C.POLICE_CORPS, size=10, bold=True, color="A16207")))

    # Journal
    ws = feuille(wb, "Journal", "Trace horodatée des actions. En cas de litige, c'est cette feuille qui parle.")
    cols = [("Date", 13), ("Heure", 10), ("Opérateur", 18), ("IP source", 18),
            ("Actif ciblé", 26), ("Action", 56), ("Résultat", 32),
            ("Autorisée par le RoE", 18)]
    d = tableau(ws, 5, cols, "Journal", 200)
    liste(ws, "H", ["Oui", "Hors RoE - justifié", "Hors RoE - incident"], d, d + 199)

    wb.save(chemin)
    print("écrit", chemin)


# --- Pilotage société --------------------------------------------------------
def classeur_pilotage(chemin: str) -> None:
    wb = Workbook()

    ws = feuille(wb, "Portefeuille", "Toutes les missions, de la prospection à la clôture.", True)
    cols = [("Référence", 22), ("Client", 22), ("Service", 22), ("Statut", 14),
            ("Chef de mission", 18), ("Montant", 14), ("Signature", 13),
            ("Début", 13), ("Fin prévue", 13), ("Livraison", 13),
            ("Destruction preuves", 17), ("Doctrine", 16), ("Notes", 34)]
    d = tableau(ws, 5, cols, "Portefeuille", 60)
    liste(ws, "C", ["pentest-audit", "ai-redteaming", "secu-applicative", "devsecops",
                    "soc-ai-tools", "x-privacy", "sensibilisation",
                    "infra-vpn-cloudflare"], d, d + 59)
    liste(ws, "D", ["prospect", "cadrage", "signé", "en-cours", "livré", "retest", "clos"],
          d, d + 59)

    ws = feuille(wb, "Plan de charge", "Disponibilité de l'équipe par mois. Sert à dire non à temps.")
    mois = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Août", "Sep", "Oct", "Nov", "Déc"]
    cols = [("Membre", 22), ("Rôle principal", 24), ("Jours ouvrés/mois", 16)] + \
           [(m, 8) for m in mois] + [("Total engagé", 13), ("Taux de charge", 14)]
    d = tableau(ws, 5, cols, "PlanCharge", 15)
    for r in range(d, d + 15):
        ws.cell(r, 16, f"=SUM(D{r}:O{r})").font = F_CORPS
        ws.cell(r, 17, f"=IF(C{r}=\"\",\"\",SUM(D{r}:O{r})/(C{r}*12))").font = F_CORPS
        ws.cell(r, 17).number_format = "0%"
    ws.conditional_formatting.add(
        f"Q{d}:Q{d + 14}",
        CellIsRule(operator="greaterThan", formula=["0.85"],
                   fill=PatternFill("solid", fgColor="FEE2E2"),
                   font=Font(name=C.POLICE_CORPS, size=10, bold=True, color="991B1B")))
    ws.cell(d + 17, 1, "Au-delà de 85 % de charge, on ne prend plus de mission : "
                       "la marge restante absorbe les imprévus et la R&D.").font = F_AIDE

    ws = feuille(wb, "Compétences", "Matrice de compétences. Identifie les points de fragilité à un seul homme.")
    domaines = ["Pentest web", "Pentest interne / AD", "Mobile", "Cloud", "AI RedTeaming",
                "DevSecOps / CI-CD", "SIEM / détection", "Forensic / DFIR", "CTI",
                "GRC / ISO 27001", "RGPD / vie privée", "Rédaction de rapport",
                "Relation client", "Avant-vente"]
    cols = [("Domaine", 28)] + [(f"Membre {i}", 14) for i in range(1, 8)] + \
           [("Couverture", 12), ("Risque", 24)]
    d = tableau(ws, 5, cols, "Competences", len(domaines) + 4)
    for i, dom in enumerate(domaines):
        ws.cell(d + i, 1, dom).font = F_CORPS
        ws.cell(d + i, 9, f"=COUNTIF(B{d + i}:H{d + i},\">=3\")").font = F_CORPS
        ws.cell(d + i, 10, f'=IF(I{d + i}=0,"Aucune couverture",'
                           f'IF(I{d + i}=1,"Dépendance à une seule personne","Couvert"))').font = F_CORPS
    liste(ws, "B", ["0", "1", "2", "3", "4"], d, d + len(domaines) - 1,
          "0 = aucune notion · 1 = notions · 2 = autonome accompagné · "
          "3 = autonome · 4 = référent, capable de former")
    for col in "CDEFGH":
        liste(ws, col, ["0", "1", "2", "3", "4"], d, d + len(domaines) - 1)
    ws.conditional_formatting.add(
        f"J{d}:J{d + len(domaines) - 1}",
        CellIsRule(operator="equal", formula=['"Aucune couverture"'],
                   fill=PatternFill("solid", fgColor="991B1B"),
                   font=Font(name=C.POLICE_CORPS, size=10, bold=True, color=C.BLANC)))
    ws.conditional_formatting.add(
        f"J{d}:J{d + len(domaines) - 1}",
        CellIsRule(operator="equal", formula=['"Dépendance à une seule personne"'],
                   fill=PatternFill("solid", fgColor="FEF3C7"),
                   font=Font(name=C.POLICE_CORPS, size=10, bold=True, color="A16207")))

    ws = feuille(wb, "Grille tarifaire", "Tarifs jour/homme. Se relit chaque année.")
    cols = [("Service", 26), ("Prestation", 40), ("Unité", 14),
            ("Junior", 12), ("Confirmé", 12), ("Senior", 12), ("Expert", 12),
            ("Durée type", 14), ("Commentaire", 34)]
    tableau(ws, 5, cols, "Tarifs", 40)

    wb.save(chemin)
    print("écrit", chemin)


# --- SoA ISO 27001 -----------------------------------------------------------
def classeur_soa(chemin: str) -> None:
    wb = Workbook()
    ws = feuille(wb, "SoA", "Déclaration d'applicabilité ISO/IEC 27001:2022 - Annexe A, 93 mesures.", True)
    themes = [("5", "Mesures organisationnelles", 37),
              ("6", "Mesures liées aux personnes", 8),
              ("7", "Mesures physiques", 14),
              ("8", "Mesures technologiques", 34)]
    cols = [("Mesure", 12), ("Thème", 26), ("Intitulé", 52), ("Applicable", 13),
            ("Justification si exclue", 40), ("Statut", 16), ("Responsable", 18),
            ("Preuve / document", 34), ("Échéance", 13)]
    total = sum(n for _, _, n in themes)
    d = tableau(ws, 5, cols, "SoA", total + 5,
                aide="Les intitulés officiels des 93 mesures sont à recopier depuis la norme "
                     "(document sous droits - ne pas le versionner dans le dépôt).")
    r = d
    for num, nom, nb in themes:
        for i in range(1, nb + 1):
            ws.cell(r, 1, f"A.{num}.{i}").font = F_CORPS
            ws.cell(r, 2, nom).font = F_CORPS
            r += 1
    liste(ws, "D", ["Oui", "Non"], d, d + total - 1)
    liste(ws, "F", ["Non commencé", "En cours", "Mis en œuvre", "Vérifié"], d, d + total - 1)
    ws.conditional_formatting.add(
        f"F{d}:F{d + total - 1}",
        CellIsRule(operator="equal", formula=['"Vérifié"'],
                   fill=PatternFill("solid", fgColor="15803D"),
                   font=Font(name=C.POLICE_CORPS, size=10, bold=True, color=C.BLANC)))
    wb.save(chemin)
    print("écrit", chemin)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="90-templates/build")
    a = ap.parse_args()
    os.makedirs(a.out, exist_ok=True)
    classeur_mission(os.path.join(a.out, "Classeur-mission.xlsx"))
    classeur_pilotage(os.path.join(a.out, "Pilotage-societe.xlsx"))
    classeur_soa(os.path.join(a.out, "SoA-ISO27001.xlsx"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
